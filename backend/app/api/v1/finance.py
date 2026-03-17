import os
import uuid as uuid_module
import csv
import io
from fastapi import APIRouter, Depends, HTTPException, Request, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from typing import List, Optional
from uuid import UUID
from datetime import date, datetime
from decimal import Decimal
from app.db.session import get_db
from app.models.user import User
from app.models.finance import Transaction, Budget, TransactionType
from app.models.ziswaf import ZiswafTransaction
from app.models.jamaah import ZakatFitrah, Jamaah
from app.schemas.finance import (
    IncomeCreate,
    ExpenseCreate,
    TransactionCreate,
    TransactionUpdate,
    TransactionResponse,
    BudgetCreate,
    BudgetUpdate,
    BudgetResponse,
    FinancialSummary,
)
from app.api.v1.auth import require_module
from app.services import cache

router = APIRouter()

UPLOADS_DIR = "uploads/receipts"


# ─── Transaction endpoints ───────────────────────────────────────────────────

@router.get("/transactions", response_model=List[TransactionResponse])
async def list_transactions(
    request: Request,
    skip: int = 0,
    limit: int = 100,
    transaction_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    date_only: Optional[date] = Query(None, alias="date"),  # shorthand for single day
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """List all transactions with optional filters."""
    if date_only:
        start_date = date_only
        end_date = date_only

    params = {
        "skip": skip, "limit": limit, "transaction_type": transaction_type,
        "start_date": str(start_date), "end_date": str(end_date),
    }
    cache_key = cache.make_key(cache._get_client_ip(request), "list_transactions", params)
    cached = await cache.get(cache_key)
    if cached is not None:
        return cached

    query = db.query(Transaction)

    if transaction_type:
        query = query.filter(Transaction.transaction_type == transaction_type)

    if start_date:
        query = query.filter(Transaction.transaction_date >= start_date)

    if end_date:
        query = query.filter(Transaction.transaction_date <= end_date)

    transactions = (
        query
        .order_by(Transaction.transaction_date.desc(), Transaction.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    result = [TransactionResponse.model_validate(t) for t in transactions]
    await cache.set(cache_key, [r.model_dump() for r in result])
    return result


@router.post("/transactions", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_transaction(
    payload: TransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Create a new transaction (unified endpoint for income and expense)."""
    transaction_data = payload.model_dump()
    transaction_data["created_by"] = current_user.id

    db_transaction = Transaction(**transaction_data)
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    await cache.queue_invalidation()
    return TransactionResponse.model_validate(db_transaction)


@router.post("/transactions/income", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_income(
    income: IncomeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Create a new income transaction."""
    transaction_data = income.model_dump()
    transaction_data["created_by"] = current_user.id

    db_transaction = Transaction(**transaction_data)
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    await cache.queue_invalidation()
    return TransactionResponse.model_validate(db_transaction)


@router.post("/transactions/expense", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_expense(
    expense: ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Create a new expense transaction."""
    transaction_data = expense.model_dump()
    transaction_data["created_by"] = current_user.id

    db_transaction = Transaction(**transaction_data)
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    await cache.queue_invalidation()
    return TransactionResponse.model_validate(db_transaction)


@router.get("/transactions/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Get a specific transaction by ID."""
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transaction not found",
        )
    return TransactionResponse.model_validate(transaction)


@router.put("/transactions/{transaction_id}", response_model=TransactionResponse)
async def update_transaction(
    transaction_id: UUID,
    transaction_update: TransactionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Update a transaction."""
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transaction not found",
        )

    update_data = transaction_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(transaction, field, value)

    db.commit()
    db.refresh(transaction)
    await cache.queue_invalidation()
    return TransactionResponse.model_validate(transaction)


@router.delete("/transactions/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction(
    transaction_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Delete a transaction."""
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transaction not found",
        )

    db.delete(transaction)
    db.commit()
    await cache.queue_invalidation()
    return None


@router.post("/transactions/{transaction_id}/approve", response_model=TransactionResponse)
async def approve_transaction(
    transaction_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Approve a transaction."""
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transaction not found",
        )

    transaction.approved_by = current_user.id
    transaction.approved_at = datetime.utcnow()

    db.commit()
    db.refresh(transaction)
    return TransactionResponse.model_validate(transaction)


# ─── Financial summary endpoint ──────────────────────────────────────────────

@router.get("/finance/summary", response_model=FinancialSummary)
async def get_financial_summary(
    request: Request,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Get financial summary for a period."""
    if not start_date:
        now = datetime.now()
        start_date = date(now.year, now.month, 1)

    if not end_date:
        end_date = date.today()

    cache_key = cache.make_key(
        cache._get_client_ip(request), "finance_summary",
        {"start_date": str(start_date), "end_date": str(end_date)},
    )
    cached = await cache.get(cache_key)
    if cached is not None:
        return cached

    income_sum = db.query(func.sum(Transaction.amount)).filter(
        and_(
            Transaction.transaction_type == TransactionType.INCOME,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
        )
    ).scalar() or Decimal(0)

    expense_sum = db.query(func.sum(Transaction.amount)).filter(
        and_(
            Transaction.transaction_type == TransactionType.EXPENSE,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
        )
    ).scalar() or Decimal(0)

    result = FinancialSummary(
        total_income=income_sum,
        total_expense=expense_sum,
        balance=income_sum - expense_sum,
        period_start=start_date,
        period_end=end_date,
    )
    await cache.set(cache_key, result.model_dump())
    return result


# ─── Receipt upload ───────────────────────────────────────────────────────────

@router.post("/finance/upload-receipt")
async def upload_receipt(
    file: UploadFile = File(...),
    current_user: User = Depends(require_module("finance")),
):
    """Upload a receipt file and return its URL."""
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'jpg'
    filename = f"{uuid_module.uuid4()}.{ext}"
    path = os.path.join(UPLOADS_DIR, filename)
    with open(path, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"url": f"/uploads/receipts/{filename}"}


# ─── Export endpoint ──────────────────────────────────────────────────────────

@router.get("/finance/export")
async def export_finance(
    format: str = "csv",  # csv, excel, pdf
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Export transactions as CSV, Excel, or PDF."""
    if not start_date:
        now = datetime.now()
        start_date = date(now.year, now.month, 1)
    if not end_date:
        end_date = date.today()

    transactions = (
        db.query(Transaction)
        .filter(
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date,
        )
        .order_by(Transaction.transaction_date.asc())
        .all()
    )

    rows = []
    for t in transactions:
        rows.append({
            "Tanggal": str(t.transaction_date),
            "Tipe": "Pemasukan" if str(t.transaction_type) in ("income", "TransactionType.INCOME") else "Pengeluaran",
            "Kategori": t.income_category or t.expense_category or "-",
            "Deskripsi": t.description or "",
            "Jumlah": float(t.amount),
            "Metode": t.payment_method or "cash",
            "Donatur/Vendor": t.donor_name or t.vendor_name or "",
            "No. Referensi": t.reference_number or "",
        })

    fieldnames = ["Tanggal", "Tipe", "Kategori", "Deskripsi", "Jumlah", "Metode", "Donatur/Vendor", "No. Referensi"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        output.seek(0)
        filename = f"mutasi-{start_date}-sd-{end_date}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    elif format == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Mutasi Keuangan"
            headers = fieldnames
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = PatternFill("solid", fgColor="4CAF50")
                cell.font = Font(bold=True, color="FFFFFF")
            for row_idx, row in enumerate(rows, 2):
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=row[key])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"mutasi-{start_date}-sd-{end_date}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
        except ImportError:
            return {"error": "openpyxl not installed"}

    elif format == "pdf":
        try:
            from fpdf import FPDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "Mutasi Keuangan Masjid", ln=True, align="C")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 8, f"Periode: {start_date} s.d. {end_date}", ln=True, align="C")
            pdf.ln(5)
            # Header row
            pdf.set_fill_color(76, 175, 80)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 9)
            cols = [("Tanggal", 28), ("Tipe", 24), ("Kategori", 30), ("Deskripsi", 55), ("Jumlah", 35)]
            for col, w in cols:
                pdf.cell(w, 8, col, border=1, fill=True)
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 8)
            total_income = 0
            total_expense = 0
            for row in rows:
                fill = row["Tipe"] == "Pemasukan"
                if fill:
                    pdf.set_fill_color(240, 255, 240)
                else:
                    pdf.set_fill_color(255, 240, 240)
                pdf.cell(28, 7, str(row["Tanggal"]), border=1, fill=True)
                pdf.cell(24, 7, str(row["Tipe"]), border=1, fill=True)
                pdf.cell(30, 7, str(row["Kategori"])[:15], border=1, fill=True)
                pdf.cell(55, 7, str(row["Deskripsi"])[:30], border=1, fill=True)
                jumlah = float(row["Jumlah"])
                sign = "+" if row["Tipe"] == "Pemasukan" else "-"
                pdf.cell(35, 7, f"{sign}Rp {jumlah:,.0f}", border=1, fill=True)
                pdf.ln()
                if row["Tipe"] == "Pemasukan":
                    total_income += jumlah
                else:
                    total_expense += jumlah
            # Summary
            pdf.ln(3)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 8, f"Total Pemasukan: Rp {total_income:,.0f}", ln=True)
            pdf.cell(0, 8, f"Total Pengeluaran: Rp {total_expense:,.0f}", ln=True)
            pdf.cell(0, 8, f"Saldo: Rp {total_income - total_expense:,.0f}", ln=True)
            output = io.BytesIO(pdf.output())
            filename = f"mutasi-{start_date}-sd-{end_date}.pdf"
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
        except ImportError:
            return {"error": "fpdf2 not installed"}

    return {"error": "Format tidak dikenal"}


# ─── Backfill endpoint ────────────────────────────────────────────────────────

TYPE_TO_CATEGORY = {
    "zakat_mal": "zakat_mal",
    "zakat_profesi": "zakat_profesi",
    "infaq": "infaq",
    "shadaqah": "shadaqah",
    "wakaf_tunai": "wakaf_tunai",
    "wakaf_aset": "wakaf_aset",
}


@router.post("/finance/backfill")
def backfill_finance(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Sync existing ZISWAF and Zakat Fitrah records that are missing finance entries."""
    created = 0

    # Backfill ZISWAF
    ziswaf_records = db.query(ZiswafTransaction).all()
    for z in ziswaf_records:
        already = db.query(Transaction).filter(
            Transaction.source_type == "ziswaf",
            Transaction.source_id == z.id,
        ).first()
        if already:
            continue
        type_val = str(z.type).replace("ZiswafType.", "")
        amount = getattr(z, "asset_value", None) if type_val == "wakaf_aset" else None
        if not amount:
            amount = z.amount
        if not amount or amount <= 0:
            continue
        trx = Transaction(
            transaction_type=TransactionType.INCOME,
            amount=amount,
            transaction_date=z.transaction_date,
            income_category=TYPE_TO_CATEGORY.get(type_val, "lainnya"),
            donor_name=z.donor_name,
            donor_jamaah_id=z.jamaah_id,
            is_anonymous=z.is_anonymous or False,
            description=f"ZISWAF - {type_val.replace('_', ' ').title()}",
            payment_method="cash",
            source_type="ziswaf",
            source_id=z.id,
            created_by=current_user.id,
        )
        db.add(trx)
        created += 1

    # Backfill Zakat Fitrah (paid only, uang > 0)
    zakat_records = db.query(ZakatFitrah).filter(
        ZakatFitrah.is_paid == True,
        ZakatFitrah.amount_uang > 0,
    ).all()
    for r in zakat_records:
        already = db.query(Transaction).filter(
            Transaction.source_type == "zakat_fitrah",
            Transaction.source_id == r.id,
        ).first()
        if already:
            continue
        donor_name = None
        if r.jamaah_id:
            j = db.get(Jamaah, r.jamaah_id)
            donor_name = j.full_name if j else None
        paid_date = r.paid_at.date() if r.paid_at else date.today()
        trx = Transaction(
            transaction_type=TransactionType.INCOME,
            amount=r.amount_uang,
            transaction_date=paid_date,
            income_category="zakat_fitrah",
            donor_name=donor_name,
            donor_jamaah_id=r.jamaah_id,
            description=f"Zakat Fitrah - {donor_name or 'Jamaah'} ({r.jumlah_jiwa} jiwa)",
            payment_method="cash",
            source_type="zakat_fitrah",
            source_id=r.id,
            created_by=current_user.id,
        )
        db.add(trx)
        created += 1

    db.commit()
    return {"synced": created, "message": f"Berhasil sinkronisasi {created} transaksi"}


# ─── Budget endpoints ─────────────────────────────────────────────────────────

@router.get("/budgets", response_model=List[BudgetResponse])
async def list_budgets(
    year: Optional[str] = None,
    month: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """List all budgets with optional filters."""
    query = db.query(Budget)

    if year:
        query = query.filter(Budget.year == year)

    if month:
        query = query.filter(Budget.month == month)

    budgets = query.offset(skip).limit(limit).all()
    return [BudgetResponse.model_validate(b) for b in budgets]


@router.post("/budgets", response_model=BudgetResponse, status_code=status.HTTP_201_CREATED)
async def create_budget(
    budget: BudgetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("finance")),
):
    """Create a new budget."""
    budget_data = budget.model_dump()
    budget_data["created_by"] = current_user.id

    db_budget = Budget(**budget_data)
    db.add(db_budget)
    db.commit()
    db.refresh(db_budget)
    return BudgetResponse.model_validate(db_budget)
